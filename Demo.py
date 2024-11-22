from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from datetime import datetime

def init_workbook(template_file):
    """
    Mở file Excel template và tạo sheet mới dựa trên template.
    """
    workbook = openpyxl.load_workbook(template_file)
    template_sheet = workbook['template']

    now = datetime.now()
    new_sheet_name = f"So sanh {now.strftime('%d-%m %H-%M')}"  # Định dạng hợp lệ
    new_sheet = workbook.copy_worksheet(template_sheet)
    new_sheet.title = new_sheet_name

    return workbook, new_sheet

def save_workbook(workbook, filename):
    """
    Lưu file Excel.
    """
    workbook.save(filename)
    print(f"Đã lưu dữ liệu vào file Excel: {filename}")

def open_tgdd_page():
    """
    Mở trang Thegioididong và đặt trình duyệt ở nửa màn hình bên trái.
    """
    chrome_options = Options()
    driver = webdriver.Chrome(options=chrome_options)

    screen_width = driver.execute_script("return window.screen.width")
    screen_height = driver.execute_script("return window.screen.height")

    driver.get("https://www.thegioididong.com")
    driver.set_window_position(0, 0)
    driver.set_window_size(screen_width // 2, screen_height)

    return driver

def get_data_tgdd(driver, sheet, start_row):
    """
    Lấy dữ liệu từ Thegioididong và ghi vào sheet Excel.
    """
    driver.find_element(By.ID, 'skw').send_keys('iphone 16 promax')
    driver.find_element(By.XPATH, "//button[i[@class='icon-search']]").click()

    WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.XPATH, "//li[contains(@class, 'item cat42')]"))
    )

    products = driver.find_elements(By.XPATH, "//li[contains(@class, 'item cat42')]")

    for row_index, product in enumerate(products, start=start_row):
        try:
            product_name = product.find_element(By.XPATH, ".//h3").text.strip()
            product_price = product.find_element(By.XPATH, ".//strong[@class='price']").text.strip()
            product_url = product.find_element(By.XPATH, ".//a").get_attribute("href")

            sheet.cell(row=row_index, column=1, value=product_name)
            sheet.cell(row=row_index, column=2, value=product_price)
            sheet.cell(row=row_index, column=3, value=product_url)

            print(f"Đã ghi TGDD: {product_name}, {product_price}, {product_url}")
        except Exception as e:
            print(f"Lỗi khi xử lý sản phẩm TGDD: {e}")

def open_cellphone_page():
    """
    Mở trang CellphoneS và đặt trình duyệt ở nửa màn hình bên phải.
    """
    chrome_options = Options()
    driver = webdriver.Chrome(options=chrome_options)

    screen_width = driver.execute_script("return window.screen.width")
    screen_height = driver.execute_script("return window.screen.height")

    driver.get("https://cellphones.com.vn")
    driver.set_window_position(screen_width // 2, 0)
    driver.set_window_size(screen_width // 2, screen_height)

    return driver

def get_data_cellphone(driver, sheet, start_row):
    """
    Lấy dữ liệu từ CellphoneS và ghi vào sheet Excel.
    """
    driver.find_element(By.ID, 'inp$earch').send_keys('iphone 16 promax')
    driver.find_element(By.XPATH, "//div[@class='input-group-btn']").click()

    WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.XPATH, '//div[@class="product-info-container product-item"]'))
    )

    products = driver.find_elements(By.XPATH, '//div[@class="product-info-container product-item"]')

    start_column = 6
    for row_index, product in enumerate(products, start=start_row):
        try:
            product_name = product.find_element(By.XPATH, ".//div[@class='product__name']/h3").text.strip()
            product_price = product.find_element(By.XPATH, './/p[@class="product__price--show"]').text.strip()
            product_url = product.find_element(By.XPATH, ".//a[@class='product__link button__link']").get_attribute("href")

            sheet.cell(row=row_index, column=start_column, value=product_name)
            sheet.cell(row=row_index, column=start_column + 1, value=product_price)
            sheet.cell(row=row_index, column=start_column + 2, value=product_url)

            print(f"Đã ghi CellphoneS: {product_name}, {product_price}, {product_url}")
        except Exception as e:
            print(f"Lỗi khi xử lý sản phẩm CellphoneS: {e}")
